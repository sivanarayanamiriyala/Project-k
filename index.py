from flask import Flask,render_template,request,session,redirect,url_for,jsonify,flash
import mysql.connector
import datetime
from flask import Response
from datetime import date
import os
from openpyxl import load_workbook
from functools import wraps


app=Flask(__name__)
app.secret_key = "sivanarayana"  

mydb = mysql.connector.connect(host="localhost",user="root",password="siva1234",database="KL")	

def login_required(f):
	@wraps(f)
	def decorated_function(*args, **kwargs):
		if 'login' not in session or session['login']!=True:
			return redirect(url_for('homepage'))
		return f(*args, **kwargs)
	return decorated_function


def operator_login(f):
	@wraps(f)
	def decorated_function(*args, **kwargs):
		if 'login' in session  and session['role']=='operator':
			return f(*args, **kwargs)
		else:
			return redirect(url_for('homepage'))
	return decorated_function


def supervisor_login(f):
	@wraps(f)
	def decorated_function(*args, **kwargs):
		if 'login' in session  and session['role']=='supervisor':
			return f(*args, **kwargs)
		else:
			return redirect(url_for('homepage'))
	return decorated_function




def manager_login(f):
	@wraps(f)
	def decorated_function(*args, **kwargs):
		if 'login' in session  and session['role']=='manager':
			return f(*args, **kwargs)
		else:
			return redirect(url_for('homepage'))
	return decorated_function





@app.route('/')
def homepage():
	[session.pop(key) for key in list(session.keys()) if not key=='_flashes']
	return render_template('login.html')

@app.route('/validate',methods=['POST'])
def validate():
	if request.method=="POST":
		username=request.form['username']
		designation=request.form['designation']
		password=request.form['password']
		cursor1=mydb.cursor(buffered=True)	
		cursor1.execute("SELECT name,empcode from users where username=%s and password=%s and designation=%s",[username,password,designation])
		data=cursor1.fetchone()
		cursor1.close()
		try:
			if data[0] and data[1]:
				name=data[0]
				empcode=data[1]
				if(designation=="manager"):
					session['login']=True
					session['name']=name
					session['role']='manager'
					session['empcode']=empcode
					return redirect(url_for('manager'))
				elif(designation=="supervisor"):
					session['login']=True
					session['name']=name
					session['role']='supervisor'
					session['empcode']=empcode
					return redirect(url_for('supervisor'))

				elif designation=="operator":
					session['login']=True
					session['name']=name
					session['role']='operator'
					session['empcode']=empcode
					return redirect(url_for('operator'))
				else:
					flash("Invalid Credentials!!")
					return redirect(url_for('homepage'))
			else:
				flash("Invalid Credentials!!")
				return redirect(url_for('homepage'))
		except Exception as e:
				flash("Invalid Credentials!!")
				return redirect(url_for('homepage'))
			




@app.route('/manager')
@login_required
@manager_login
def manager():
	return render_template('managerhome.html')

@app.route('/assignoperator')
@login_required
@manager_login
def assignoperator():
	cursor1=mydb.cursor(buffered=True)	
	cursor1.execute("SELECT empcode,name from users where designation='supervisor' and isworking=1")
	supervisor=cursor1.fetchall()
	cursor1.close()

	cursor1=mydb.cursor(buffered=True)	
	cursor1.execute("SELECT empcode,name from users where designation='operator' and isworking=1")
	operator=cursor1.fetchall()
	cursor1.close()
	return render_template("assignemployees.html",supervisor=supervisor,operator=operator)


@app.route('/storeoperatorassignment',methods=["POST"])
@login_required
def storeoperatorassignment():
	if request.method=="POST":
		supervisor=request.form['supervisor']
		empcode=request.form['empcode']

		cursor3=mydb.cursor()	
		cursor3.execute("update users set supervisor=%s where empcode=%s",(supervisor,empcode))
		mydb.commit()
		flash("Operator Assigned to the Supervisor")
		return redirect(url_for('assignoperator'))


@app.route('/viewassignedoperators')
@login_required
@manager_login
def viewassignedoperators():
	cursor1=mydb.cursor(buffered=True)	
	cursor1.execute("SELECT empcode,name from users where designation='supervisor' and isworking=1")
	supervisor=cursor1.fetchall()
	cursor1.close()
	mydb.commit()

	return render_template('viewassignedoperators.html',supervisor=supervisor)

@app.route('/showassignedoperator',methods=["POST"])
@login_required
def showassignedoperator():
	if request.method=="POST":
		supervisor=request.form['supervisor']
		cursor1=mydb.cursor(buffered=True)	
		cursor1.execute("SELECT t2.empcode,t2.name,t1.empcode,t1.name  FROM users t1 JOIN users t2 ON t1.supervisor = t2.name WHERE t1.supervisor = %s and t1.isworking=1 and t2.isworking=1",[supervisor])
		data=cursor1.fetchall()
		cursor1.close()
		return jsonify({'data':data})

@app.route('/addemployee')
@login_required
@manager_login
def addemployee():
	cursor1=mydb.cursor(buffered=True)	
	cursor1.execute("SELECT name from users where designation='supervisor' and isworking=1")
	name=cursor1.fetchall()
	cursor1.close()
	return render_template('addemployee.html',supervisornames=name)

@app.route('/modify')
@manager_login
@login_required
def modify():
	cursor1=mydb.cursor(buffered=True)	
	cursor1.execute("SELECT name from users where designation='supervisor' and isworking=1")
	name=cursor1.fetchall()
	cursor1.close()
	return render_template('modify.html',supervisornames=name)


@app.route('/getuserinformation',methods=['POST'])
def getuserinformation():
	if request.method=="POST":
		empcode=request.form['enteredempcode']
		cursor1=mydb.cursor(buffered=True)	
		cursor1.execute("SELECT * from users where empcode=%s and isworking=1",[empcode])
		data=cursor1.fetchone()
		cursor1.close()
		return jsonify({'data':data})



@app.route('/modifyemployeedata',methods=["POST"])
@login_required
def modifyemployeedata():
	if request.method=="POST":
		empcode=request.form['empcode']
		username=request.form['username']
		name=request.form['name']
		password=request.form['password']
		designation=request.form['designation']
		cursor3=mydb.cursor()
		try:
			cursor3.execute("update users set username=%s,name=%s,password=%s,designation=%s where empcode=%s and isworking=1",(username,name,password,designation,empcode))
			mydb.commit()
			flash("Data Modified")
			mydb.commit()
		except Exception as e:
			print(e)			

		return redirect(url_for('modify'))



@app.route('/deleteemployee',methods=["POST"])
@login_required
def deleteemployee():
	if request.method=="POST":
		empcode=request.form['empcode']
		cursor3=mydb.cursor()
		try:
			cursor3.execute("update users set isworking=0 where empcode=%s",[empcode])
			mydb.commit()
			return jsonify({'data':'Employee Deleted'})

		except Exception as e:
			print(e)	
			return jsonify({'data':'Error'})



@app.route('/storeemployee',methods=["POST"])
@login_required
def storeemployee():
	if request.method=="POST":
		empcode=request.form['empcode']
		username=request.form['username']
		name=request.form['name']
		password=request.form['password']
		designation=request.form['designation']

		cursor3=mydb.cursor()
		try:
			cursor3.execute("insert into users(empcode,username,name,password,designation) values(%s,%s,%s,%s,%s)",(empcode,username,name,password,designation))
			mydb.commit()
			flash("Employee Added")
		except Exception as e:
			print(e)			
		return redirect(url_for('addemployee'))

@app.route('/storemachine',methods=["POST"])
@login_required
def storemachine():
	if request.method=="POST":
		machinecode=request.form['machinecode']
		cursor3=mydb.cursor()	
		cursor3.execute("insert into machine values(%s)",(machinecode,))
		mydb.commit()
		flash("Machine Added")
		return redirect(url_for('addmachine'))



@app.route('/getuserinfo',methods=['POST'])
def getuserinfo():
	if request.method=="POST":
		username=request.form['username']
		designation=request.form['designation']
		cursor1=mydb.cursor(buffered=True)	
		cursor1.execute("SELECT name,supervisor from users where username=%s and designation=%s and isworking=1",[username,designation])
		data=cursor1.fetchone()
		cursor1.close()
		return jsonify({'data':data})


@app.route('/data')
@login_required
@manager_login
def data():
	return render_template('data.html')



@app.route('/addmachine')
@login_required
@manager_login
def addmachine():
	return render_template('addmachine.html')



@app.route('/operator')
@login_required
@operator_login
def operator():
	cursor1=mydb.cursor(buffered=True)	
	cursor1.execute("SELECT hours from managercredentials")
	hours=cursor1.fetchone()
	cursor1.close()

	return render_template('operator.html',hours=hours[0])


@app.route('/getdata',methods=['POST'])
@login_required
def getdata():
	if request.method=="POST":
		empcode=request.form['empcode']
		date=datetime.datetime.now().date()

		cursor1=mydb.cursor(buffered=True)
		cursor1.execute("select machinecode,shift from assignedduties where empcode=%s",[empcode]) 
		data9=cursor1.fetchone()
		cursor1.close()
		if data9:
			shift=data9[1]

			cursor1=mydb.cursor(buffered=True)	
			cursor1.execute("SELECT name from users where empcode=%s",(empcode,))
			data=cursor1.fetchall()
			cursor1.close()
		
			cursor1=mydb.cursor(buffered=True)
			cursor1.execute("select modelname,operation,assempcode,hourlytype,heatno,produced,passed,reject,lossno,lossinmin,materialrejectiontype,materialrejectioncount,processrejectiontype,processrejctioncount,hasloss from hourlydata where hour=1 and empcode=%s and date=%s and shift=%s",(empcode,date,shift,)) 
			data1=cursor1.fetchall()
			cursor1.close()
			
			cursor1=mydb.cursor(buffered=True)
			cursor1.execute("select modelname,operation,assempcode,hourlytype,heatno,produced,passed,reject,lossno,lossinmin,materialrejectiontype,materialrejectioncount,processrejectiontype,processrejctioncount,hasloss from hourlydata where hour=2 and empcode=%s and date=%s and shift=%s",(empcode,date,shift,)) 
			data2=cursor1.fetchall()
			cursor1.close()

			cursor1=mydb.cursor(buffered=True)
			cursor1.execute("select modelname,operation,assempcode,hourlytype,heatno,produced,passed,reject,lossno,lossinmin,materialrejectiontype,materialrejectioncount,processrejectiontype,processrejctioncount,hasloss from hourlydata where hour=3 and empcode=%s and date=%s and shift=%s",(empcode,date,shift,)) 
			data3=cursor1.fetchall()
			cursor1.close()

			cursor1=mydb.cursor(buffered=True)
			cursor1.execute("select modelname,operation,assempcode,hourlytype,heatno,produced,passed,reject,lossno,lossinmin,materialrejectiontype,materialrejectioncount,processrejectiontype,processrejctioncount,hasloss from hourlydata where hour=4 and empcode=%s and date=%s and shift=%s",(empcode,date,shift,)) 
			data4=cursor1.fetchall()
			cursor1.close()

			cursor1=mydb.cursor(buffered=True)
			cursor1.execute("select modelname,operation,assempcode,hourlytype,heatno,produced,passed,reject,lossno,lossinmin,materialrejectiontype,materialrejectioncount,processrejectiontype,processrejctioncount,hasloss from hourlydata where hour=5 and empcode=%s and date=%s and shift=%s",(empcode,date,shift,)) 
			data5=cursor1.fetchall()
			cursor1.close()

			cursor1=mydb.cursor(buffered=True)
			cursor1.execute("select modelname,operation,assempcode,hourlytype,heatno,produced,passed,reject,lossno,lossinmin,materialrejectiontype,materialrejectioncount,processrejectiontype,processrejctioncount,hasloss from hourlydata where hour=6 and empcode=%s and date=%s and shift=%s",(empcode,date,shift,)) 
			data6=cursor1.fetchall()
			cursor1.close()

			cursor1=mydb.cursor(buffered=True)
			cursor1.execute("select modelname,operation,assempcode,hourlytype,heatno,produced,passed,reject,lossno,lossinmin,materialrejectiontype,materialrejectioncount,processrejectiontype,processrejctioncount,hasloss from hourlydata where hour=7 and empcode=%s and date=%s and shift=%s",(empcode,date,shift,)) 
			data7=cursor1.fetchall()
			cursor1.close()
	
			cursor1=mydb.cursor(buffered=True)
			cursor1.execute("select modelname,operation,assempcode,hourlytype,heatno,produced,passed,reject,lossno,lossinmin,materialrejectiontype,materialrejectioncount,processrejectiontype,processrejctioncount,hasloss from hourlydata where hour=8 and empcode=%s and date=%s and shift=%s",(empcode,date,shift,)) 
			data8=cursor1.fetchall()
			cursor1.close()

			cursor1=mydb.cursor(buffered=True)
			cursor1.execute("select * from losses") 
			data10=cursor1.fetchall()
			cursor1.close()

			return jsonify({'data':data[0],'data1':data1,'data2':data2,'data3':data3,'data4':data4,'data5':data5,'data6':data6,'data7':data7,'data8':data8,'data9':data9,'data10':data10})

		else:
			cursor1=mydb.cursor(buffered=True)	
			cursor1.execute("SELECT * from users where empcode=%s",[empcode])
			data=cursor1.fetchall()
			cursor1.close()
			if data:		
				return jsonify({'output':"Didn't Assigned any Machine"})
			else:
				return jsonify({'output':"Invalid Employee Code"})


@app.route('/datastore',methods=['POST'])
@login_required
def datastore():
	if request.method=="POST":
		shift=request.form['shift']
		hour=request.form['hour']
		heatno=request.form['heatno']
		produce=request.form['produce']
		reject=request.form['reject']
		lossno=request.form['lossno']
		passed=request.form['passed']
		lossinmin=request.form['lossinmin']
		machinecode=request.form['machinecode']
		empcode=request.form['empcode']

		assempcode=request.form['assempcode']
		hourlytype=request.form['hourlytype']

		date=datetime.datetime.now().date()

		modelname=request.form['modelname']
		operation=request.form['operation']
		empname=request.form['empname']
		materialrejectiontype=request.form['materialrejectiontype']
		materialrejectioncount=request.form['materialrejectioncount']
		processrejectioncount=request.form['processrejectioncount']


		totalmaterialrejectioncount=request.form['totalmaterialrejectioncount']
		totalprocessrejectioncount=request.form['totalprocessrejectioncount']
		totalproduce=request.form['totalproduce']
		totalreject=request.form['totalreject']
		totalpassed=request.form['totalpassed']
		totallossinmin=request.form['totallossinmin']
		
		if not materialrejectioncount:
			materialrejectioncount=0
		if not processrejectioncount:
			processrejectioncount=0

		processrejectiontype=request.form['processrejectiontype']
		hasloss=request.form['hasloss']

		cursor3=mydb.cursor()	
		cursor3.execute("insert into hourlydata values(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",(empcode,empname,machinecode,modelname,assempcode,operation,date,shift,hourlytype,hour,heatno,produce,passed,reject,materialrejectiontype,materialrejectioncount,processrejectiontype,processrejectioncount,hasloss,lossno,lossinmin,))
		mydb.commit()
		
		cursor3=mydb.cursor()	
		cursor3.execute("select supervisor from users where empcode=%s",[empcode])
		supervisor=cursor3.fetchone()
		mydb.commit()


		cursor3=mydb.cursor()	


		cursor3.execute("INSERT INTO dailydata(empcode,empname,machinecode,supervisor,modelname,assempcode,operation,date,shift,hourlytype,produced,passed, reject,materialrejectioncount,processrejectioncount,lossinmin) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) ON DUPLICATE KEY UPDATE produced=%s,passed=%s,reject=%s,materialrejectioncount=%s,processrejectioncount=%s,lossinmin=%s",[empcode,empname,machinecode,supervisor[0],modelname,assempcode,operation,date,shift,hourlytype,totalproduce,totalpassed,totalreject,totalmaterialrejectioncount,totalprocessrejectioncount,totallossinmin,totalproduce,totalpassed,totalreject,totalmaterialrejectioncount,totalprocessrejectioncount,totallossinmin])
		mydb.commit()

		return jsonify({'success':'success'})




@app.route('/supervisor')
@login_required
@supervisor_login
def supervisor():
	return render_template('supervisorhome.html')



@app.route('/assignmachine')
@login_required
@supervisor_login
def assignmachine():
	cursor1=mydb.cursor(buffered=True)	

	cursor1.execute("SELECT empcode,name FROM users WHERE designation='operator' AND isworking=1 AND supervisor=(SELECT name FROM users WHERE empcode=%s)",[session['empcode']])
	operator=cursor1.fetchall()
	cursor1.close()


	cursor1=mydb.cursor(buffered=True)	
	cursor1.execute("SELECT machinecode from machine")
	machine=cursor1.fetchall()
	cursor1.close()


	cursor1=mydb.cursor(buffered=True)	

	cursor1.execute("SELECT shift from managercredentials")
	shift=cursor1.fetchone()
	cursor1.close()

	return render_template("assignmachines.html",operator=operator,machine=machine,shift=shift[0])




@app.route('/storemachineassignment',methods=["POST"])
@login_required
@supervisor_login
def storemachineassignment():
	if request.method=="POST":
		machinecode=request.form['machinecode']
		empcode=request.form['empcode']
		shift=request.form['shift']

		cursor3=mydb.cursor()
		cursor3.execute("INSERT INTO assignedduties(empcode,machinecode,shift) VALUES(%s,%s,%s) ON DUPLICATE KEY UPDATE machinecode=%s,shift=%s",[empcode,machinecode,shift,machinecode,shift])
		mydb.commit()
		flash("Machine  Assigned")
		return redirect(url_for('assignmachine'))



@app.route('/generatereport')
@login_required
@supervisor_login
def generatereport():
	cursor1=mydb.cursor(buffered=True)	

	cursor1.execute("SELECT shift from managercredentials")
	shift=cursor1.fetchone()
	cursor1.close()

	return render_template("generatereport.html",shift=shift[0])


@app.route('/generate',methods=["POST"])
@login_required
def generate():
	if request.method=="POST":
		date=request.form['date']
		shift=request.form['shift']
		report=request.form['report']
		global data
		if(report=="hourlydata"):
			cursor=mydb.cursor()	
	
			cursor.execute("""
SELECT
    empcode,
    machinecode,
	modelname,
operation,
    MAX(CASE WHEN hour = 1 THEN heatno END) AS heatno_1,
    MAX(CASE WHEN hour = 1 THEN produced END) AS produced_1,
    MAX(CASE WHEN hour = 1 THEN passed END) AS passed_1,
    MAX(CASE WHEN hour = 1 THEN reject END) AS reject_1,
    MAX(CASE WHEN hour = 1 THEN lossno END) AS lossno_1,
    MAX(CASE WHEN hour = 1 THEN lossinmin END) AS lossinmin_1,


    MAX(CASE WHEN hour = 2 THEN heatno END) AS heatno_2,
    MAX(CASE WHEN hour = 2 THEN produced END) AS produced_2,
    MAX(CASE WHEN hour = 2 THEN passed END) AS passed_2,
    MAX(CASE WHEN hour = 2 THEN reject END) AS reject_2,
    MAX(CASE WHEN hour = 2 THEN lossno END) AS lossno_2,
    MAX(CASE WHEN hour = 2 THEN lossinmin END) AS lossinmin_2,

    MAX(CASE WHEN hour = 3 THEN heatno END) AS heatno_3,
    MAX(CASE WHEN hour = 3 THEN produced END) AS produced_3,
    MAX(CASE WHEN hour = 3 THEN passed END) AS passed_3,
    MAX(CASE WHEN hour = 3 THEN reject END) AS reject_3,
    MAX(CASE WHEN hour = 3 THEN lossno END) AS lossno_3,
    MAX(CASE WHEN hour = 3 THEN lossinmin END) AS lossinmin_3,


    MAX(CASE WHEN hour = 4 THEN heatno END) AS heatno_4,
    MAX(CASE WHEN hour = 4 THEN produced END) AS produced_4,
    MAX(CASE WHEN hour = 4 THEN passed END) AS passed_4,
    MAX(CASE WHEN hour = 4 THEN reject END) AS reject_4,
    MAX(CASE WHEN hour = 4 THEN lossno END) AS lossno_4,
    MAX(CASE WHEN hour = 4 THEN lossinmin END) AS lossinmin_4,


    MAX(CASE WHEN hour = 5 THEN heatno END) AS heatno_5,
    MAX(CASE WHEN hour = 5 THEN produced END) AS produced_5,
    MAX(CASE WHEN hour = 5 THEN passed END) AS passed_5,
    MAX(CASE WHEN hour = 5 THEN reject END) AS reject_5,
    MAX(CASE WHEN hour = 5 THEN lossno END) AS lossno_5,
    MAX(CASE WHEN hour = 5 THEN lossinmin END) AS lossinmin_5,


    MAX(CASE WHEN hour = 6 THEN heatno END) AS heatno_6,
    MAX(CASE WHEN hour = 6 THEN produced END) AS produced_6,
    MAX(CASE WHEN hour = 6 THEN passed END) AS passed_6,
    MAX(CASE WHEN hour = 6 THEN reject END) AS reject_6,
    MAX(CASE WHEN hour = 6 THEN lossno END) AS lossno_6,
    MAX(CASE WHEN hour = 6 THEN lossinmin END) AS lossinmin_6,

    MAX(CASE WHEN hour = 7 THEN heatno END) AS heatno_7,
    MAX(CASE WHEN hour = 7 THEN produced END) AS produced_7,
    MAX(CASE WHEN hour = 7 THEN passed END) AS passed_7,
    MAX(CASE WHEN hour = 7 THEN reject END) AS reject_7,
    MAX(CASE WHEN hour = 7 THEN lossno END) AS lossno_7,
    MAX(CASE WHEN hour = 7 THEN lossinmin END) AS lossinmin_7,

    MAX(CASE WHEN hour = 8 THEN heatno END) AS heatno_8,
    MAX(CASE WHEN hour = 8 THEN produced END) AS produced_8,
    MAX(CASE WHEN hour = 8 THEN passed END) AS passed_8,
    MAX(CASE WHEN hour = 8 THEN reject END) AS reject_8,
    MAX(CASE WHEN hour = 8 THEN lossno END) AS lossno_8,
    MAX(CASE WHEN hour = 8 THEN lossinmin END) AS lossinmin_8

FROM hourlydata
WHERE hour BETWEEN 1 AND 8
    AND shift = %s
    AND date = %s
GROUP BY empcode,machinecode,modelname,operation
""",(shift,date));
			data=cursor.fetchall()
	
			mydb.commit()
			wb = load_workbook('static/report.xlsx')
			wb1 = load_workbook('static/test.xlsx')
			try:	
				del wb1['Sheet1']
			except:
				pass	
			wb.save('test.xlsx')
			work_sheet = wb.active
			for i in data:
				work_sheet.append(i)
		elif report=="dailydata":
			cursor=mydb.cursor()	
	
			cursor.execute("SELECT * from dailydata where shift=%s and date=%s",(shift,date));
			data=cursor.fetchall()
			mydb.commit()
			wb = load_workbook('static/report1.xlsx')
			wb1 = load_workbook('static/test1.xlsx')
			try:	
				del wb1['Sheet1']
			except:
				pass	
			wb.save('test1.xlsx')
			work_sheet = wb.active
			for i in data:
				work_sheet.append(i)
		filename='test1.xlsx'
		wb.save('static/test1.xlsx')


		return jsonify({'output':'success','data':data})




@app.route('/viewreportgenerate',methods=["POST"])
@login_required
def viewreportgenerate():
	if request.method=="POST":
		date=request.form['date']
		shift=request.form['shift']
		cursor=mydb.cursor()	
		cursor.execute("SELECT * FROM hourlydata WHERE  shift = %s AND date = %s AND empcode IN (SELECT empcode FROM users WHERE supervisor=(SELECT name FROM users WHERE empcode = %s ))",[shift,date,session['empcode']])
		data=cursor.fetchall()
		mydb.commit()
		return jsonify({'data':data})


@app.route('/viewassignedmachines')
@login_required
@supervisor_login
def viewassignedmachines():
	cursor1=mydb.cursor(buffered=True)	
	cursor1.execute("SELECT machinecode from machine")
	machinecode=cursor1.fetchall()
	cursor1.close()
	mydb.commit()
	return render_template('viewassignedmachines.html',machinecode=machinecode)


@app.route('/showassignedmachines',methods=["POST"])
@login_required
def showassignedmachines():
	if request.method=="POST":
		machinecode=request.form['machinecode']
		cursor1=mydb.cursor(buffered=True)	
		if machinecode!="all":
			cursor1.execute("SELECT machinecode,empcode,(select name from users where empcode=assignedduties.empcode),shift  FROM assignedduties where machinecode=%s and assignedduties.empcode in (select empcode from users where supervisor=%s )",[machinecode,session['name']])
		else:
			cursor1.execute("SELECT machinecode,empcode,(select name from users where empcode=assignedduties.empcode),shift  FROM assignedduties where assignedduties.empcode in (select empcode from users where supervisor=%s )",[session['name']])
		data=cursor1.fetchall()
		cursor1.close()
		return jsonify({'data':data})


@app.route('/forgotpassword')
def forgotpassword():
	return "Contact Your Manager to Reset the PassWord"



@app.route('/viewdailyreportgenerate',methods=["POST"])
@login_required
def viewdailyreportgenerate():
	if request.method=="POST":
		date=request.form['date']
		shift=request.form['shift']
		cursor=mydb.cursor()	
		cursor.execute("SELECT * FROM dailydata WHERE  shift = %s AND date = %s AND empcode IN (SELECT empcode FROM users WHERE supervisor=(SELECT name FROM users WHERE empcode = %s ) and isworking=1)",[shift,date,session['empcode']])
		data=cursor.fetchall()
		mydb.commit()
		return jsonify({'data':data})



@app.route('/viewmanagerreport',methods=["POST"])
@login_required
def viewmanagerreport():
	if request.method=="POST":
		to=request.form['to']
		fromdate=request.form['from']
		cursor=mydb.cursor()	
		cursor.execute("SELECT supervisor,sum(produced),sum(passed),sum(reject),sum(materialrejectioncount),sum(processrejectioncount),sum(lossinmin) FROM dailydata WHERE date between %s AND %s group by supervisor",[fromdate,to])
		data=cursor.fetchall()
		mydb.commit()
		return jsonify({'data':data})


@app.route('/changeshiftdetails')
@login_required
@manager_login
def changeshiftdetails():

	cursor=mydb.cursor()	
	cursor.execute("SELECT * from managercredentials")
	data=cursor.fetchall()
	mydb.commit()
	return render_template('changeshiftdetails.html',data=data)


@app.route('/changeshiftinfo',methods=['POST'])
@login_required
def changeshiftinfo():
	if request.method=="POST":
		shift=request.form['shift']
		hours=request.form['hours']
		cursor=mydb.cursor()	
		cursor.execute("update managercredentials set shift=%s, hours=%s",[shift,hours])
		mydb.commit()
		flash('Changes Updated')
		return redirect('changeshiftdetails')




@app.route('/getoperatordetails',methods=['POST'])
def getoperatordetails():
	if  request.method=="POST":
		operatorcode=request.form['operatorcode']
		cursor=mydb.cursor()
		cursor.execute("select machinecode,shift from assignedduties where empcode=%s",[operatorcode])
		data=cursor.fetchone()
		mydb.commit()
		return jsonify({'data':data})


@app.route('/getshiftdetails',methods=['POST'])
def getshiftdetails():
	if  request.method=="POST":
		machinecode=request.form['machinecode']
		shift=request.form['shift']		
		cursor=mydb.cursor()
		cursor.execute("select empcode,(select name from users where empcode=assignedduties.empcode) from assignedduties where machinecode=%s and shift=%s",[machinecode,shift])
		data=cursor.fetchone()
		mydb.commit()
		return jsonify({'data':data})




if __name__=='__main__':
	app.run(debug=True)